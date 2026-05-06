"""
loader.py — Charge les documents sources des 5 dossiers CACEIS.

Formats supportés :
  - PDF   : pdfplumber (texte natif)
  - DOCX  : ZIP+XML (robuste, gère les fichiers complexes)
  - PPTX  : ZIP+XML (évite le bug 'rId' de python-pptx)
  - XLSX  : openpyxl (résumé texte des données structurées)

Retourne : list[dict] avec clés source, folder, content, type
"""

import os
import re
import logging
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

import pdfplumber
import openpyxl

logging.basicConfig(level=logging.INFO, format="%(levelname)s | %(message)s")
log = logging.getLogger(__name__)

# ── Dossiers sources ──────────────────────────────────────────────────────────
BASE = Path("/Users/rayanekryslak-medioub/Desktop/CACEIS/Sujet Alberthon")

FOLDERS = [
    BASE / "Employee satisfaction & engagement",
    BASE / "Finance",
    BASE / "Governance",
    BASE / "HR Data",
    BASE / "Training",
]

# Fichiers à ignorer explicitement (trop gros, données individuelles RGPD)
SKIP_FILES = {
    "Absentéisme_-_détail_affectation_-_Bilan_social.xlsx",
    "Absentéisme_-_détail_affectation_-_Bilan_social (1).xlsx",
    "20260121 - Absentéisme_-_détail_affectation_-_Bilan_social 2025.xlsx",
    "2025 - Stats CACEIS EAE EP fichier de travail - Vretraitement.xlsx",
    ".DS_Store",
}

MIN_WORDS = 80  # seuil minimum pour indexer un document


# ── Extracteurs ───────────────────────────────────────────────────────────────

def _extract_pdf(path: Path) -> str:
    """Extrait le texte d'un PDF via pdfplumber."""
    try:
        text_parts = []
        with pdfplumber.open(str(path)) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    text_parts.append(t)
        return "\n".join(text_parts)
    except Exception as e:
        log.warning(f"PDF extraction error [{path.name}]: {e}")
        return ""


def _extract_docx(path: Path) -> str:
    """Extrait le texte d'un DOCX via ZIP+XML (robuste)."""
    try:
        with zipfile.ZipFile(str(path), "r") as z:
            if "word/document.xml" not in z.namelist():
                return ""
            xml_str = z.read("word/document.xml").decode("utf-8", errors="ignore")
        root = ET.fromstring(xml_str)
        words = []
        for elem in root.iter():
            if elem.tag.endswith("}t") and elem.text and elem.text.strip():
                words.append(elem.text.strip())
        return " ".join(words)
    except Exception as e:
        log.warning(f"DOCX extraction error [{path.name}]: {e}")
        return ""


def _extract_pptx(path: Path) -> str:
    """Extrait le texte d'un PPTX via ZIP+XML (évite le bug 'rId')."""
    try:
        words = []
        with zipfile.ZipFile(str(path), "r") as z:
            slide_files = sorted(
                [n for n in z.namelist()
                 if re.match(r"ppt/slides/slide\d+\.xml", n)]
            )
            for sfile in slide_files:
                xml_str = z.read(sfile).decode("utf-8", errors="ignore")
                root = ET.fromstring(xml_str)
                for elem in root.iter():
                    if elem.tag.endswith("}t") and elem.text and elem.text.strip():
                        words.append(elem.text.strip())
        return " ".join(words)
    except Exception as e:
        log.warning(f"PPTX extraction error [{path.name}]: {e}")
        return ""


def _extract_xlsx(path: Path) -> str:
    """
    Extrait une représentation textuelle d'un XLSX.
    Lit les 3 premiers onglets, max 200 lignes par onglet.
    Produit : "Feuille NOM\ncolA | colB | ...\nval1 | val2 | ..."
    Évite toute donnée individuelle (RGPD) — résumé structurel uniquement.
    """
    try:
        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        parts = []
        for sheet_name in wb.sheetnames[:3]:
            ws = wb[sheet_name]
            rows_text = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 200:
                    break
                # Filtre les lignes vides
                cells = [str(c).strip() if c is not None else "" for c in row]
                if any(c for c in cells):
                    rows_text.append(" | ".join(cells[:12]))  # max 12 colonnes
            if rows_text:
                parts.append(f"Feuille: {sheet_name}\n" + "\n".join(rows_text))
        wb.close()
        return "\n\n".join(parts)
    except Exception as e:
        log.warning(f"XLSX extraction error [{path.name}]: {e}")
        return ""


# ── Chargeur principal ────────────────────────────────────────────────────────

def load_all_documents(folders=None) -> list[dict]:
    """
    Charge tous les documents depuis les dossiers sources.

    Returns:
        list[dict] avec clés :
          source   : nom du fichier
          folder   : nom du dossier parent
          content  : texte extrait
          type     : pdf | docx | pptx | xlsx
    """
    if folders is None:
        folders = FOLDERS

    documents = []
    skipped = []

    for folder in folders:
        folder = Path(folder)
        if not folder.exists():
            log.warning(f"Dossier introuvable : {folder}")
            continue

        folder_name = folder.name

        for file_path in sorted(folder.iterdir()):
            if not file_path.is_file():
                continue
            if file_path.name.startswith("."):
                continue
            if file_path.name in SKIP_FILES:
                log.info(f"  SKIP (liste) : {file_path.name}")
                skipped.append(file_path.name)
                continue

            ext = file_path.suffix.lower().lstrip(".")

            if ext == "pdf":
                content = _extract_pdf(file_path)
                doc_type = "pdf"
            elif ext == "docx":
                content = _extract_docx(file_path)
                doc_type = "docx"
            elif ext == "pptx":
                content = _extract_pptx(file_path)
                doc_type = "pptx"
            elif ext == "xlsx":
                # Skip fichiers très lourds (>8 MB) — données brutes individuelles
                if file_path.stat().st_size > 8 * 1024 * 1024:
                    log.info(f"  SKIP (>8MB) : {file_path.name}")
                    skipped.append(file_path.name)
                    continue
                content = _extract_xlsx(file_path)
                doc_type = "xlsx"
            elif ext == "json":
                continue  # ignoré
            else:
                log.info(f"  SKIP (type non supporté) : {file_path.name}")
                skipped.append(file_path.name)
                continue

            # Filtre contenu trop court
            word_count = len(content.split())
            if word_count < MIN_WORDS:
                log.info(f"  SKIP (<{MIN_WORDS} mots, {word_count} extraits) : {file_path.name}")
                skipped.append(file_path.name)
                continue

            documents.append({
                "source":    file_path.name,
                "folder":    folder_name,
                "content":   content,
                "type":      doc_type,
                "word_count": word_count,
            })
            log.info(f"  OK  [{doc_type.upper():4s}] {file_path.name} ({word_count} mots)")

    log.info(f"\n{'='*50}")
    log.info(f"Documents chargés : {len(documents)}")
    log.info(f"Fichiers ignorés  : {len(skipped)}")
    return documents
